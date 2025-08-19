import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import logging
from datetime import datetime
import os
from api_client import OpenAPIClient
from semantic_mapper import load_excel, extract_template_structure, match_rows, build_output

logger = logging.getLogger(__name__)

class ExcelProcessor:
    """Process Excel files and generate formatted output using OpenAI"""
    
    def __init__(self):
        self.openai_client = OpenAPIClient()
        # Use project-relative template path
        self.template_path = "document/Template-Waaree.xlsx"

    def process_with_semantic_mapping(self, uploaded_file_path):
        input_data = load_excel(uploaded_file_path)
        template_data = load_excel(self.template_path)
        template_structure = extract_template_structure(template_data)

        # Build per-sheet row mappings
        row_mapping_by_sheet = {}
        for sheet_name in template_structure:
            input_sheet = input_data.get(sheet_name)
            if input_sheet is not None and not input_sheet.empty:
                template_rows = template_structure[sheet_name]["rows"]
                input_rows = input_sheet.iloc[:, 0].dropna().astype(str).tolist()
                sheet_row_mapping = match_rows(template_rows, input_rows)
                row_mapping_by_sheet[sheet_name] = sheet_row_mapping

        processed_data = build_output(template_structure, input_data, row_mapping_by_sheet)
        return processed_data


    def read_excel_file(self, file_path):
        """
        Read Excel file and extract data
        
        Args:
            file_path (str): Path to the Excel file
            
        Returns:
            dict: Extracted data from the file
        """
        try:
            # Read all sheets
            excel_file = pd.ExcelFile(file_path)
            data = {}
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                # Convert DataFrame to records and handle NaN values
                records = df.to_dict('records')
                # Clean the data by replacing NaN with None and handling time objects
                cleaned_records = []
                for record in records:
                    cleaned_record = {}
                    for key, value in record.items():
                        if pd.isna(value):
                            cleaned_record[key] = None
                        elif hasattr(value, 'time') or hasattr(value, 'strftime'):
                            # Convert time/datetime objects to strings
                            if hasattr(value, 'strftime'):
                                cleaned_record[key] = value.strftime('%H:%M:%S')
                            else:
                                cleaned_record[key] = str(value)
                        else:
                            cleaned_record[key] = value
                    cleaned_records.append(cleaned_record)
                data[sheet_name] = cleaned_records
            
            logger.info(f"Successfully read Excel file: {file_path}")
            return data
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            raise
    
    def analyze_template(self):
        """
        Analyze the template file to understand its structure
        
        Returns:
            dict: Template structure information
        """
        try:
            template_data = self.read_excel_file(self.template_path)
            
            # Analyze structure
            structure = {
                "sheets": {},
                "headers": {},
                "data_types": {},
                "template_rows": {}
            }
            
            for sheet_name, data in template_data.items():
                if data:
                    structure["sheets"][sheet_name] = {
                        "columns": list(data[0].keys()),
                        "row_count": len(data)
                    }
                    
                    # Include actual template rows so OpenAI knows which rows to map to
                    structure["template_rows"][sheet_name] = data
                    
                    # Analyze data types and convert to JSON-serializable format
                    df = pd.DataFrame(data)
                    dtypes_dict = {}
                    for col, dtype in df.dtypes.items():
                        # Convert pandas dtype to string representation
                        dtypes_dict[str(col)] = str(dtype)
                    structure["data_types"][sheet_name] = dtypes_dict
            
            logger.info("Template analysis completed")
            return structure
            
        except Exception as e:
            logger.error(f"Error analyzing template: {str(e)}")
            raise
    
    def create_openai_prompt(self, uploaded_data, template_structure):
        """
        Create a prompt for OpenAI to process the data
        
        Args:
            uploaded_data (dict): Data from uploaded file
            template_structure (dict): Template structure information
            
        Returns:
            str: Formatted prompt for OpenAI
        """
        # Filter and limit data to prevent token limit issues
        filtered_data = self._filter_relevant_data(uploaded_data, template_structure)
        
        # Create a simplified template structure with only essential information
        simplified_template = {
            "sheets": template_structure.get("sheets", {}),
            "template_rows": {}
        }
        
        # Include key rows from each template sheet to show structure
        for sheet_name, rows in template_structure.get("template_rows", {}).items():
            if rows:
                # Include first 3 rows, last 2 rows, and a few key middle rows to show structure
                key_rows = []
                if len(rows) >= 3:
                    key_rows.extend(rows[:3])  # First 3 rows
                if len(rows) >= 5:
                    # Add a middle row if there are enough rows
                    middle_idx = len(rows) // 2
                    key_rows.append(rows[middle_idx])
                if len(rows) >= 2:
                    key_rows.extend(rows[-2:])  # Last 2 rows
                
                # Limit to maximum 10 rows to keep prompt manageable
                simplified_template["template_rows"][sheet_name] = key_rows[:10]
        
        prompt = f"""
You are an Excel data processing expert. I need you to process uploaded data and format it according to a specific template structure.

TEMPLATE STRUCTURE (SIMPLIFIED):
{json.dumps(simplified_template, indent=2)}

UPLOADED DATA (FILTERED FOR RELEVANT SHEETS):
{json.dumps(filtered_data, indent=2)}

INSTRUCTIONS:
1. Use the template structure shown above as your reference for the output format.
2. For each sheet, you must generate ALL rows that would be in the complete template (not just the sample rows shown).
3. The sample template rows show the structure - you need to create the complete set of rows for each sheet.
4. Map uploaded data to the appropriate template structure, maintaining the same column names and data types.
5. Extract relevant information from uploaded_data and populate the template structure with actual values.
6. Ensure all required fields are populated according to the template structure.
7. Maintain data integrity and formatting as defined in the template.
8. For missing data, use appropriate default values or leave as null.
9. Follow the same row structure and column names as shown in the template.
10. Generate the complete set of rows for each sheet, not just the sample rows shown.

CRITICAL OUTPUT REQUIREMENTS:
- Your output must be a JSON object with a "sheets" key
- Each sheet in "sheets" must contain an array of objects
- Each object in the array must represent a row from template_structure.template_rows
- The number of rows in your output must match exactly with template_structure.template_rows
- The order of rows must match exactly with template_structure.template_rows
- Each row object must have the same keys as defined in template_structure.template_rows

RESPONSE FORMAT:
Return a JSON object with the following structure:
{{
  "sheets": {{
    "P&L": [
      {{"Sr no.": "1", "Particulars": "Revenue", "2023-2024": 1000, "2024-2025": 1100, ...}},
      {{"Sr no.": "2", "Particulars": "Expenses", "2023-2024": 800, "2024-2025": 880, ...}},
      ...
    ],
    "BS": [
      {{"Particulars": "Assets", "2023-2024": 5000, "2024-2025": 5500, ...}},
      ...
    ],
    "Cash Flow": [
      {{"Sr No": "1", "Particulars": "Operating Cash Flow", "2023-2024": 200, "2024-2025": 220, ...}},
      ...
    ]
  }}
}}

Return only the JSON data, no additional text or explanations.
"""
        return prompt
    
    def _filter_relevant_data(self, uploaded_data, template_structure):
        """
        Filter uploaded data to include only relevant sheets and limit data size
        
        Args:
            uploaded_data (dict): Original uploaded data
            template_structure (dict): Template structure
            
        Returns:
            dict: Filtered data
        """
        template_sheets = set(template_structure.get('sheets', {}).keys())
        filtered_data = {}
        
        # Priority sheets that are most likely to contain relevant data
        priority_sheets = ['P&L', 'BS', 'Cash Flow', 'Profit & Loss', 'Balance Sheet', 'Cash Flow Statement','p&l']
        
        for sheet_name, data in uploaded_data.items():
            # Check if this sheet matches template sheets or priority sheets
            is_relevant = (
                sheet_name in template_sheets or
                any(priority in sheet_name for priority in priority_sheets) or
                any(template_sheet in sheet_name for template_sheet in template_sheets)
            )
            
            if is_relevant and data:
                # Limit the number of rows to prevent token limit issues
                max_rows = 20  # Reduced to 20 rows per sheet to prevent token limit issues
                filtered_data[sheet_name] = data[:max_rows]
                logger.info(f"Included sheet '{sheet_name}' with {len(filtered_data[sheet_name])} rows (limited from {len(data)})")
            elif data:
                logger.info(f"Skipped sheet '{sheet_name}' (not relevant to template)")
        
        # If no relevant sheets found, include first few sheets with limited data
        if not filtered_data:
            logger.warning("No relevant sheets found, including first 3 sheets with limited data")
            for i, (sheet_name, data) in enumerate(uploaded_data.items()):
                if i < 3 and data:  # Only first 3 sheets
                    filtered_data[sheet_name] = data[:10]  # Reduced to 10 rows
                    logger.info(f"Included fallback sheet '{sheet_name}' with {len(filtered_data[sheet_name])} rows")
        
        return filtered_data
    
    def process_with_openai(self, uploaded_data, template_structure):
        """
        Process uploaded data using OpenAI
        
        Args:
            uploaded_data (dict): Data from uploaded file
            template_structure (dict): Template structure information
            
        Returns:
            dict: Processed data from OpenAI
        """
        try:
            prompt = self.create_openai_prompt(uploaded_data, template_structure)
            logger.info(f"Created OpenAI prompt (length: {len(prompt)} characters)")
            
            messages = [
                {"role": "system", "content": "You are an Excel data processing expert. Process the data and return only valid JSON. Do not include any markdown formatting or code blocks."},
                {"role": "user", "content": prompt}
            ]
            
            logger.info("Sending data to OpenAI for processing")
            response = self.openai_client.make_chat_completion(
                messages=messages,
                temperature=0.1,  # Low temperature for consistent formatting
                max_tokens=16000  # Further increased for very large JSON responses
            )
            
            # Extract the response content
            content = response['choices'][0]['message']['content']
            logger.info(f"OpenAI response received (length: {len(content)} characters)")
            logger.info(f"OpenAI response preview: {content[:200]}...")
            
            # Check if response was truncated
            if response['choices'][0].get('finish_reason') == 'length':
                logger.warning("OpenAI response was truncated due to token limit")
                logger.warning("Consider reducing the amount of data sent or increasing max_tokens")
            
            # Try to parse JSON from the response
            try:
                # Clean the response to extract JSON
                content = content.strip()
                
                # Remove markdown code blocks
                if content.startswith('```json'):
                    content = content[7:]
                elif content.startswith('```'):
                    content = content[3:]
                if content.endswith('```'):
                    content = content[:-3]
                
                # Remove any leading/trailing whitespace
                content = content.strip()
                
                # Try to find JSON object boundaries if there's extra text
                start_idx = content.find('{')
                end_idx = content.rfind('}')
                
                if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                    content = content[start_idx:end_idx + 1]
                
                # Check if JSON appears to be complete
                if not content.strip().endswith('}'):
                    logger.error("JSON response appears to be incomplete (doesn't end with '}')")
                    logger.error("This might indicate a truncated response")
                    raise ValueError("Incomplete JSON response - likely truncated")
                
                logger.info(f"Cleaned response (length: {len(content)} characters)")
                logger.info(f"Cleaned response preview: {content[:200]}...")
                logger.info(f"Cleaned response end: ...{content[-200:]}")
                
                # Try to parse the JSON
                processed_data = json.loads(content)
                logger.info(f"Successfully parsed JSON. Keys: {list(processed_data.keys())}")
                
                # Validate the processed data structure
                for sheet_name, data in processed_data.items():
                    if isinstance(data, list):
                        logger.info(f"Sheet '{sheet_name}' has {len(data)} rows")
                        if data:
                            logger.info(f"First row keys: {list(data[0].keys()) if data[0] else 'No keys'}")
                    else:
                        logger.warning(f"Sheet '{sheet_name}' is not a list: {type(data)}")
                
                # Validate that processed data matches template structure
                logger.info("Validating processed data against template structure...")
                if 'sheets' in processed_data:
                    for sheet_name, processed_rows in processed_data['sheets'].items():
                        if sheet_name in template_structure.get('template_rows', {}):
                            template_rows = template_structure['template_rows'][sheet_name]
                            logger.info(f"Validating sheet '{sheet_name}': processed={len(processed_rows)} rows, template={len(template_rows)} rows")
                            
                            if len(processed_rows) != len(template_rows):
                                logger.warning(f"Row count mismatch in sheet '{sheet_name}': processed has {len(processed_rows)} rows, template has {len(template_rows)} rows")
                            
                            # Check if keys match
                            if processed_rows and template_rows:
                                processed_keys = set(processed_rows[0].keys())
                                template_keys = set(template_rows[0].keys())
                                if processed_keys != template_keys:
                                    logger.warning(f"Key mismatch in sheet '{sheet_name}': processed keys {processed_keys}, template keys {template_keys}")
                        else:
                            logger.warning(f"Sheet '{sheet_name}' not found in template structure")
                else:
                    logger.warning("Processed data does not contain 'sheets' key")
                
                logger.info("Successfully processed data with OpenAI")
                return processed_data
                
            except json.JSONDecodeError as e:
                logger.error(f"Failed to parse OpenAI response as JSON: {str(e)}")
                logger.error(f"Error position: {e.pos}")
                logger.error(f"Error line: {e.lineno}, column: {e.colno}")
                logger.error(f"Response content: {content}")
                
                # Try to show the problematic area
                if e.pos < len(content):
                    start = max(0, e.pos - 50)
                    end = min(len(content), e.pos + 50)
                    logger.error(f"Problematic area around position {e.pos}: '{content[start:end]}'")
                
                # Try to fix common JSON issues
                logger.info("Attempting to fix common JSON issues...")
                try:
                    # Try to fix trailing commas
                    import re
                    fixed_content = re.sub(r',(\s*[}\]])', r'\1', content)
                    
                    # Try to fix unquoted keys
                    fixed_content = re.sub(r'(\s*)(\w+)(\s*):', r'\1"\2"\3:', fixed_content)
                    
                    logger.info("Attempting to parse fixed JSON...")
                    processed_data = json.loads(fixed_content)
                    logger.info("Successfully parsed JSON after fixing common issues")
                    return processed_data
                    
                except json.JSONDecodeError as e2:
                    logger.error(f"Failed to parse even after fixing common issues: {str(e2)}")
                    raise ValueError("OpenAI response is not valid JSON and could not be fixed")
                
        except Exception as e:
            logger.error(f"Error processing with OpenAI: {str(e)}")
            raise
    
    def create_formatted_excel(self, processed_data, output_path):
        """
        Create formatted Excel file based on template and processed data
        
        Args:
            processed_data (dict): Processed data from OpenAI
            output_path (str): Path to save the output file
            
        Returns:
            str: Path to the created file
        """
        try:
            # Load template
            template_wb = openpyxl.load_workbook(self.template_path)
            logger.info(f"Template loaded with sheets: {template_wb.sheetnames}")
            
            # Create new workbook
            output_wb = Workbook()
            
            # Get the default sheet name
            default_sheet_name = output_wb.sheetnames[0]
            logger.info(f"Default sheet name: {default_sheet_name}")
            
            # Check if processed_data has the expected structure
            if 'sheets' not in processed_data:
                logger.error(f"Processed data does not contain 'sheets' key. Available keys: {list(processed_data.keys())}")
                raise ValueError("Processed data structure is invalid - missing 'sheets' key")
            
            # Get the actual sheet data
            sheet_data = processed_data['sheets']
            logger.info(f"Sheet data keys: {list(sheet_data.keys())}")
            
            # Copy template structure and apply data
            for sheet_name in template_wb.sheetnames:
                logger.info(f"Processing sheet: {sheet_name}")
                
                if sheet_name in sheet_data:
                    logger.info(f"Found processed data for sheet: {sheet_name}")
                    sheet_rows = sheet_data[sheet_name]
                    logger.info(f"Sheet '{sheet_name}' has {len(sheet_rows)} rows")
                    
                    # Create new sheet or use existing
                    if sheet_name in output_wb.sheetnames:
                        output_ws = output_wb[sheet_name]
                        logger.info(f"Using existing sheet: {sheet_name}")
                    else:
                        # If this is the first sheet, rename the default sheet
                        if len(output_wb.sheetnames) == 1 and output_wb.sheetnames[0] == default_sheet_name:
                            output_wb.active.title = sheet_name
                            output_ws = output_wb.active
                            logger.info(f"Renamed default sheet to: {sheet_name}")
                        else:
                            output_ws = output_wb.create_sheet(sheet_name)
                            logger.info(f"Created new sheet: {sheet_name}")
                    
                    # Copy template sheet
                    template_ws = template_wb[sheet_name]
                    logger.info(f"Template sheet dimensions: {template_ws.max_row} rows x {template_ws.max_column} columns")
                    
                    # Copy formatting and structure from template
                    for row in template_ws.iter_rows():
                        for cell in row:
                            new_cell = output_ws.cell(row=cell.row, column=cell.column)
                            new_cell.value = cell.value
                            
                            # Copy formatting
                            if cell.font:
                                new_cell.font = Font(
                                    name=cell.font.name,
                                    size=cell.font.size,
                                    bold=cell.font.bold,
                                    italic=cell.font.italic,
                                    color=cell.font.color
                                )
                            
                            if cell.alignment:
                                new_cell.alignment = Alignment(
                                    horizontal=cell.alignment.horizontal,
                                    vertical=cell.alignment.vertical,
                                    wrap_text=cell.alignment.wrap_text
                                )
                            
                            if cell.fill:
                                new_cell.fill = PatternFill(
                                    fill_type=cell.fill.fill_type,
                                    start_color=cell.fill.start_color,
                                    end_color=cell.fill.end_color
                                )
                    
                    # Apply processed data
                    if sheet_rows and len(sheet_rows) > 0:
                        logger.info(f"Applying {len(sheet_rows)} rows of data to sheet: {sheet_name}")
                        
                        # Get the column headers from the first row of processed data
                        if isinstance(sheet_rows[0], dict):
                            headers = list(sheet_rows[0].keys())
                            logger.info(f"Data headers: {headers}")
                            
                            # Write headers to first row
                            for col_idx, header in enumerate(headers, 1):
                                cell = output_ws.cell(row=1, column=col_idx)
                                cell.value = header
                                logger.info(f"Writing header '{header}' to cell {1}:{col_idx}")
                            
                            # Write data starting from row 2
                            for row_idx, row_data in enumerate(sheet_rows, 2):
                                logger.info(f"Writing row {row_idx}: {row_data}")
                                for col_idx, (col_name, value) in enumerate(row_data.items(), 1):
                                    cell = output_ws.cell(row=row_idx, column=col_idx)
                                    # Handle None values
                                    if value is None:
                                        cell.value = ""
                                    else:
                                        cell.value = value
                                    logger.info(f"Writing '{value}' to cell {row_idx}:{col_idx}")
                        else:
                            logger.warning(f"Sheet data is not in expected format. First row type: {type(sheet_rows[0])}")
                    else:
                        logger.warning(f"No processed data found for sheet: {sheet_name}")
                else:
                    logger.warning(f"No processed data found for template sheet: {sheet_name}")
            
            # Remove the default sheet only if we have other sheets
            if len(output_wb.sheetnames) > 1 and default_sheet_name in output_wb.sheetnames:
                output_wb.remove(output_wb[default_sheet_name])
                logger.info(f"Removed default sheet: {default_sheet_name}")
            
            # Ensure we have at least one sheet
            if len(output_wb.sheetnames) == 0:
                # Create a default sheet if no sheets exist
                output_wb.create_sheet("Sheet1")
                logger.warning("No sheets were created, added default Sheet1")
            
            logger.info(f"Final workbook sheets: {output_wb.sheetnames}")
            
            # Save the file
            output_wb.save(output_path)
            logger.info(f"Formatted Excel file created: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error creating formatted Excel: {str(e)}")
            logger.error(f"Exception type: {type(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise
    
    def process_uploaded_file(self, uploaded_file_path):
        """
        Main method to process uploaded Excel file
        
        Args:
            uploaded_file_path (str): Path to uploaded file
            
        Returns:
            str: Path to the generated output file
        """
        try:
            logger.info(f"Processing uploaded file: {uploaded_file_path}")
            
            # Attempt semantic/embedding-based direct mapping first to avoid LLM hallucinations
            processed_data = None
            try:
                logger.info("Attempting semantic mapping pipeline (no LLM)...")
                processed_data = self.process_with_semantic_mapping(uploaded_file_path)
                if processed_data and isinstance(processed_data, dict) and processed_data.get('sheets'):
                    logger.info("Semantic mapping succeeded. Skipping LLM processing.")
                else:
                    logger.warning("Semantic mapping returned empty or invalid data. Will fall back to LLM.")
                    processed_data = None
            except Exception as sem_err:
                logger.warning(f"Semantic mapping failed, will fall back to LLM. Reason: {sem_err}")

            if processed_data is None:
                # Read uploaded file (for LLM path)
                uploaded_data = self.read_excel_file(uploaded_file_path)
                logger.info(f"Uploaded data sheets: {list(uploaded_data.keys())}")
                for sheet_name, data in uploaded_data.items():
                    logger.info(f"Sheet '{sheet_name}' has {len(data)} rows")
                    if data:
                        logger.info(f"Sample data from '{sheet_name}': {data[0] if data else 'No data'}")
                
                # Analyze template
                template_structure = self.analyze_template()
                logger.info(f"Template structure: {json.dumps(template_structure, indent=2)}")
                
                # Process with OpenAI (fallback)
                logger.info("Starting OpenAI processing (fallback)...")
                processed_data = self.process_with_openai(uploaded_data, template_structure)
                logger.info(f"OpenAI processing completed. Processed data keys: {list(processed_data.keys())}")
            
            # Log processed data structure
            if 'sheets' in processed_data:
                sheet_data = processed_data['sheets']
                for sheet_name, data in sheet_data.items():
                    if isinstance(data, list):
                        logger.info(f"Processed sheet '{sheet_name}' has {len(data)} rows")
                        if data:
                            logger.info(f"Processed data sample from '{sheet_name}': {data[0] if data else 'No data'}")
                    else:
                        logger.info(f"Processed sheet '{sheet_name}' has structure: {type(data)}")
            else:
                logger.warning("No 'sheets' key found in processed data")
            
            # Generate output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"processed_{timestamp}.xlsx"
            output_path = os.path.join("document", output_filename)
            
            # Create formatted Excel file
            logger.info("Creating formatted Excel file...")
            os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
            self.create_formatted_excel(processed_data, output_path)
            
            logger.info(f"File processing completed: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error processing uploaded file: {str(e)}")
            logger.error(f"Exception type: {type(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise 