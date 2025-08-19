import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json
import logging
from datetime import datetime
import os
import PyPDF2
import re
from api_client import OpenAPIClient

logger = logging.getLogger(__name__)


class FileProcessor:
    """Process various file types (Excel, PDF) and generate formatted output using OpenAI"""

    def __init__(self):
        self.openai_client = OpenAPIClient()
        self.template_path = "document/Template-Waaree.xlsx"

    def detect_file_type(self, file_path):
        """
        Detect the type of file based on extension and content

        Args:
            file_path (str): Path to the file

        Returns:
            str: File type ('excel', 'pdf', 'unknown')
        """
        file_extension = file_path.lower().split('.')[-1]

        if file_extension in ['xlsx', 'xls']:
            return 'excel'
        elif file_extension == 'pdf':
            return 'pdf'
        else:
            return 'unknown'

    def read_excel_file(self, file_path):
        """
        Read Excel file and extract data

        Args:
            file_path (str): Path to the Excel file

        Returns:
            dict: Extracted data from the file
        """
        try:
            excel_file = pd.ExcelFile(file_path)
            data = {}

            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                records = df.to_dict('records')

                cleaned_records = []
                for record in records:
                    cleaned_record = {}
                    for key, value in record.items():
                        if pd.isna(value):
                            cleaned_record[key] = None
                        elif hasattr(value, 'strftime'):
                            cleaned_record[key] = value.strftime('%H:%M:%S')
                        else:
                            cleaned_record[key] = value
                    cleaned_records.append(cleaned_record)
                data[sheet_name] = cleaned_records

            logger.info(f"Successfully read Excel file: {file_path}")
            return data

        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            raise

    def read_pdf_file(self, file_path):
        """
        Read PDF file and extract text data

        Args:
            file_path (str): Path to the PDF file

        Returns:
            dict: Extracted text data from the file
        """
        try:
            data = {}

            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)

                all_text = ""
                for page_num, page in enumerate(pdf_reader.pages):
                    page_text = page.extract_text()
                    all_text += f"\n--- Page {page_num + 1} ---\n{page_text}"

                sections = self._parse_financial_sections(all_text)

                for section_name, section_data in sections.items():
                    data[section_name] = section_data

            logger.info(f"Successfully read PDF file: {file_path}")
            return data

        except Exception as e:
            logger.error(f"Error reading PDF file: {str(e)}")
            raise

    def _parse_financial_sections(self, text):
        """
        Parse financial sections from PDF text

        Args:
            text (str): Extracted text from PDF

        Returns:
            dict: Parsed financial sections
        """
        sections = {}

        patterns = {
            'P&L': r'(?:profit\s*&\s*loss|income\s*statement|p&l|revenue|income|expenses)',
            'BS': r'(?:balance\s*sheet|assets|liabilities|equity)',
            'Cash Flow': r'(?:cash\s*flow|cash\s*and\s*cash\s*equivalents|operating\s*activities)',
            'Notes': r'(?:notes|disclosures|accounting\s*policies)'
        }

        lines = text.split('\n')
        current_section = 'General'
        current_data = []

        for line in lines:
            line = line.strip()
            if not line:
                continue

            section_found = False
            for section_name, pattern in patterns.items():
                if re.search(pattern, line, re.IGNORECASE):
                    if current_data:
                        sections[current_section] = current_data
                    current_section = section_name
                    current_data = []
                    section_found = True
                    break

            if not section_found:
                parsed_line = self._parse_financial_line(line)
                if parsed_line:
                    current_data.append(parsed_line)

        if current_data:
            sections[current_section] = current_data

        return sections

    def _parse_financial_line(self, line):
        """
        Parse a line of text to extract financial data

        Args:
            line (str): Line of text

        Returns:
            dict: Parsed financial data or None
        """
        line = re.sub(r'\s+', ' ', line.strip())

        if re.search(r'[0-9,]+\.?[0-9]*', line):
            parts = re.split(r'[:|]\s*', line)
            if len(parts) >= 2:
                key = parts[0].strip()
                value = parts[1].strip()
                return {key: value}

        return None

    def read_file(self, file_path):
        """
        Read file based on its type

        Args:
            file_path (str): Path to the file

        Returns:
            dict: Extracted data from the file
        """
        file_type = self.detect_file_type(file_path)

        if file_type == 'excel':
            return self.read_excel_file(file_path)
        elif file_type == 'pdf':
            return self.read_pdf_file(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

    def analyze_template(self):
        """
        Analyze the template file to understand its structure

        Returns:
            dict: Template structure information
        """
        try:
            template_data = self.read_excel_file(self.template_path)

            structure = {
                "sheets": {},
                "headers": {},
                "data_types": {},
                "template_rows": {}
            }

            for sheet_name, data in template_data.items():
                if data:
                    structure["sheets"][sheet_name] = {
                        "columns": list(data[0].keys()),
                        "row_count": len(data)
                    }

                    structure["template_rows"][sheet_name] = data

                    df = pd.DataFrame(data)
                    dtypes_dict = {}
                    for col, dtype in df.dtypes.items():
                        dtypes_dict[str(col)] = str(dtype)
                    structure["data_types"][sheet_name] = dtypes_dict

            logger.info("Template analysis completed")
            return structure

        except Exception as e:
            logger.error(f"Error analyzing template: {str(e)}")
            raise

    def create_openai_prompt(self, uploaded_data, template_structure, file_type):
        """
        Create a prompt for OpenAI to process the data

        Args:
            uploaded_data (dict): Data from uploaded file
            template_structure (dict): Template structure information
            file_type (str): Type of uploaded file

        Returns:
            str: Formatted prompt for OpenAI
        """
        filtered_data = self._filter_relevant_data(uploaded_data, template_structure)

        uploaded_overview = {}
        for sheet_name, rows in uploaded_data.items():
            if not rows:
                continue
            columns = list(rows[0].keys())
            first_col = columns[0] if columns else None
            row_labels = []
            if first_col:
                for rec in rows:
                    val = rec.get(first_col)
                    if val is not None and str(val).strip() != "":
                        row_labels.append(str(val))
            uploaded_overview[sheet_name] = {
                "columns": columns,
                "row_labels": row_labels,
                "sample_rows": rows[:5]
            }

        simplified_template = {
            "sheets": template_structure.get("sheets", {}),
            "template_rows": {}
        }

        for sheet_name, rows in template_structure.get("template_rows", {}).items():
            if rows:
                key_rows = []
                if len(rows) >= 3:
                    key_rows.extend(rows[:3])
                if len(rows) >= 5:
                    middle_idx = len(rows) // 2
                    key_rows.append(rows[middle_idx])
                if len(rows) >= 2:
                    key_rows.extend(rows[-2:])
                simplified_template["template_rows"][sheet_name] = key_rows[:10]

        file_type_instructions = ""
        if file_type == 'pdf':
            file_type_instructions = """
SPECIAL INSTRUCTIONS FOR PDF PROCESSING:
- The uploaded data is extracted from a PDF file and may be less structured than Excel data
- Look for financial terms and numbers in the text data
- Map financial items to appropriate template rows based on meaning and context
- Extract numerical values and financial terms from the text
- Handle cases where data might be in different formats or sections
"""

        prompt = f"""
You are an expert financial data processor. I need you to process uploaded data from a {file_type.upper()} file and format it according to a specific template structure.

{file_type_instructions}

TEMPLATE STRUCTURE (SIMPLIFIED):
{json.dumps(simplified_template, indent=2)}

UPLOADED DATA OVERVIEW (FULL ROW LABELS + COLUMNS):
{json.dumps(uploaded_overview, indent=2)}

SAMPLED ROW DATA (LIMITED TO CONTROL TOKENS):
{json.dumps(filtered_data, indent=2)}

INSTRUCTIONS:
1. Use the template structure shown above as your reference for the output format.
2. For each sheet, you must generate ALL rows that would be in the complete template (not just the sample rows shown).
3. The sample template rows show the structure - you need to create the complete set of rows for each sheet.
4. Map uploaded data to the appropriate template structure, maintaining the same column names and data types.
5. Extract relevant information from uploaded_data and populate the template structure with actual values.
6. Ensure all required fields are populated according to the template structure.
7. Maintain data integrity and formatting as defined in the template.
8. For missing data, use appropriate default values or leave as null.
9. Follow the same row structure and column names as shown in the template.
10. Generate the complete set of rows for each sheet, not just the sample rows shown.

CRITICAL OUTPUT REQUIREMENTS:
- Your output must be a JSON object with a "sheets" key
- Each sheet in "sheets" must contain an array of objects
- Each object in the array must represent a row from template_structure.template_rows
- The number of rows in your output must match exactly with template_structure.template_rows
- The order of rows must match exactly with template_structure.template_rows
- Each row object must have the same keys as defined in template_structure.template_rows

RESPONSE FORMAT:
Return a JSON object with the following structure:
{{
  "sheets": {{
    "P&L": [
      {{"Sr no.": "1", "Particulars": "Revenue", "2023-2024": 1000, "2024-2025": 1100, ...}},
      {{"Sr no.": "2", "Particulars": "Expenses", "2023-2024": 800, "2024-2025": 880, ...}},
      ...
    ],
    "BS": [
      {{"Particulars": "Assets", "2023-2024": 5000, "2024-2025": 5500, ...}},
      ...
    ],
    "Cash Flow": [
      {{"Sr No": "1", "Particulars": "Operating Cash Flow", "2023-2024": 200, "2024-2025": 220, ...}},
      ...
    ]
  }}
}}

Return only the JSON data, no additional text or explanations.
"""
        return prompt

    def _filter_relevant_data(self, uploaded_data, template_structure):
        """
        Filter uploaded data to include only relevant sheets and limit data size

        Args:
            uploaded_data (dict): Original uploaded data
            template_structure (dict): Template structure

        Returns:
            dict: Filtered data
        """
        template_sheets = set(template_structure.get('sheets', {}).keys())
        filtered_data = {}

        priority_sheets = ['P&L', 'BS', 'Cash Flow', 'Profit & Loss', 'Balance Sheet', 'Cash Flow Statement', 'p&l', 'General']

        for sheet_name, data in uploaded_data.items():
            is_relevant = (
                sheet_name in template_sheets or
                any(priority in sheet_name for priority in priority_sheets) or
                any(template_sheet in sheet_name for template_sheet in template_sheets)
            )

            if is_relevant and data:
                max_rows = 20
                filtered_data[sheet_name] = data[:max_rows]
                logger.info(f"Included sheet '{sheet_name}' with {len(filtered_data[sheet_name])} rows (limited from {len(data)})")
            elif data:
                logger.info(f"Skipped sheet '{sheet_name}' (not relevant to template)")

        if not filtered_data:
            logger.warning("No relevant sheets found, including first 3 sheets with limited data")
            for i, (sheet_name, data) in enumerate(uploaded_data.items()):
                if i < 3 and data:
                    filtered_data[sheet_name] = data[:10]
                    logger.info(f"Included fallback sheet '{sheet_name}' with {len(filtered_data[sheet_name])} rows")

        return filtered_data

    def process_with_openai(self, uploaded_data, template_structure, file_type):
        """
        Process uploaded data using OpenAI

        Args:
            uploaded_data (dict): Data from uploaded file
            template_structure (dict): Template structure information
            file_type (str): Type of uploaded file

        Returns:
            dict: Processed data from OpenAI
        """
        try:
            prompt = self.create_openai_prompt(uploaded_data, template_structure, file_type)
            logger.info(f"Created OpenAI prompt (length: {len(prompt)} characters)")

            messages = [
                {"role": "system", "content": "You are an expert financial data processor. Process the data and return only valid JSON. Do not include any markdown formatting or code blocks."},
                {"role": "user", "content": prompt}
            ]

            logger.info("Sending data to OpenAI for processing")
            response = self.openai_client.make_chat_completion(
                messages=messages,
                temperature=0.1,
                max_tokens=16000
            )

            content = response['choices'][0]['message']['content']
            logger.info(f"OpenAI response received (length: {len(content)} characters)")

            if response['choices'][0].get('finish_reason') == 'length':
                logger.warning("OpenAI response was truncated due to token limit")
                logger.warning("Consider reducing the amount of data sent or increasing max_tokens")

            try:
                content = content.strip()

                if content.startswith('```json'):
                    content = content[7:]
                elif content.startswith('```'):
                    content = content[3:]
                if content.endswith('```'):
                    content = content[:-3]

                content = content.strip()
                start_idx = content.find('{')
                end_idx = content.rfind('}')

                if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                    content = content[start_idx:end_idx + 1]

                if not content.strip().endswith('}'):
                    logger.error("JSON response appears to be incomplete (doesn't end with '}')")
                    raise ValueError("Incomplete JSON response - likely truncated")

                processed_data = json.loads(content)
                logger.info(f"Successfully parsed JSON. Keys: {list(processed_data.keys())}")

                logger.info("Validating processed data against template structure...")
                if 'sheets' in processed_data:
                    for sheet_name, processed_rows in processed_data['sheets'].items():
                        if sheet_name in template_structure.get('template_rows', {}):
                            template_rows = template_structure['template_rows'][sheet_name]
                            logger.info(f"Validating sheet '{sheet_name}': processed={len(processed_rows)} rows, template={len(template_rows)} rows")
                            if processed_rows and template_rows:
                                processed_keys = set(processed_rows[0].keys())
                                template_keys = set(template_rows[0].keys())
                                if processed_keys != template_keys:
                                    logger.warning(f"Key mismatch in sheet '{sheet_name}': processed keys {processed_keys}, template keys {template_keys}")
                else:
                    logger.warning("Processed data does not contain 'sheets' key")

                return processed_data

            except json.JSONDecodeError as e:
                logger.error(f"Failed to parse OpenAI response as JSON: {str(e)}")
                try:
                    import re as _re
                    fixed_content = _re.sub(r',(\s*[}\]])', r'\1', content)
                    fixed_content = _re.sub(r'(\s*)(\w+)(\s*):', r'\1"\2"\3:', fixed_content)
                    processed_data = json.loads(fixed_content)
                    logger.info("Successfully parsed JSON after fixing common issues")
                    return processed_data
                except json.JSONDecodeError as e2:
                    logger.error(f"Failed to parse even after fixing common issues: {str(e2)}")
                    raise ValueError("OpenAI response is not valid JSON and could not be fixed")

        except Exception as e:
            logger.error(f"Error processing with OpenAI: {str(e)}")
            raise

    def create_formatted_excel(self, processed_data, output_path):
        """
        Create formatted Excel file based on template and processed data

        Args:
            processed_data (dict): Processed data from OpenAI
            output_path (str): Path to save the output file

        Returns:
            str: Path to the created file
        """
        try:
            template_wb = openpyxl.load_workbook(self.template_path)
            logger.info(f"Template loaded with sheets: {template_wb.sheetnames}")

            output_wb = Workbook()
            default_sheet_name = output_wb.sheetnames[0]
            logger.info(f"Default sheet name: {default_sheet_name}")

            if 'sheets' not in processed_data:
                logger.error(f"Processed data does not contain 'sheets' key. Available keys: {list(processed_data.keys())}")
                raise ValueError("Processed data structure is invalid - missing 'sheets' key")

            sheet_data = processed_data['sheets']
            logger.info(f"Sheet data keys: {list(sheet_data.keys())}")

            for sheet_name in template_wb.sheetnames:
                logger.info(f"Processing sheet: {sheet_name}")

                if sheet_name in sheet_data:
                    logger.info(f"Found processed data for sheet: {sheet_name}")
                    sheet_rows = sheet_data[sheet_name]
                    logger.info(f"Sheet '{sheet_name}' has {len(sheet_rows)} rows")

                    if sheet_name in output_wb.sheetnames:
                        output_ws = output_wb[sheet_name]
                        logger.info(f"Using existing sheet: {sheet_name}")
                    else:
                        if len(output_wb.sheetnames) == 1 and output_wb.sheetnames[0] == default_sheet_name:
                            output_wb.active.title = sheet_name
                            output_ws = output_wb.active
                            logger.info(f"Renamed default sheet to: {sheet_name}")
                        else:
                            output_ws = output_wb.create_sheet(sheet_name)
                            logger.info(f"Created new sheet: {sheet_name}")

                    template_ws = template_wb[sheet_name]
                    logger.info(f"Template sheet dimensions: {template_ws.max_row} rows x {template_ws.max_column} columns")

                    for row in template_ws.iter_rows():
                        for cell in row:
                            new_cell = output_ws.cell(row=cell.row, column=cell.column)
                            new_cell.value = cell.value

                            if cell.font:
                                new_cell.font = Font(
                                    name=cell.font.name,
                                    size=cell.font.size,
                                    bold=cell.font.bold,
                                    italic=cell.font.italic,
                                    color=cell.font.color
                                )

                            if cell.alignment:
                                new_cell.alignment = Alignment(
                                    horizontal=cell.alignment.horizontal,
                                    vertical=cell.alignment.vertical,
                                    wrap_text=cell.alignment.wrap_text
                                )

                            if cell.fill:
                                new_cell.fill = PatternFill(
                                    fill_type=cell.fill.fill_type,
                                    start_color=cell.fill.start_color,
                                    end_color=cell.fill.end_color
                                )

                    if sheet_rows and len(sheet_rows) > 0:
                        logger.info(f"Applying {len(sheet_rows)} rows of data to sheet: {sheet_name}")

                        if isinstance(sheet_rows[0], dict):
                            headers = list(sheet_rows[0].keys())
                            logger.info(f"Data headers: {headers}")

                            for col_idx, header in enumerate(headers, 1):
                                cell = output_ws.cell(row=1, column=col_idx)
                                cell.value = header

                            for row_idx, row_data in enumerate(sheet_rows, 2):
                                for col_idx, (col_name, value) in enumerate(row_data.items(), 1):
                                    cell = output_ws.cell(row=row_idx, column=col_idx)
                                    cell.value = "" if value is None else value
                        else:
                            logger.warning(f"Sheet data is not in expected format. First row type: {type(sheet_rows[0])}")
                    else:
                        logger.warning(f"No processed data found for sheet: {sheet_name}")
                else:
                    logger.warning(f"No processed data found for template sheet: {sheet_name}")

            if len(output_wb.sheetnames) > 1 and default_sheet_name in output_wb.sheetnames:
                output_wb.remove(output_wb[default_sheet_name])

            if len(output_wb.sheetnames) == 0:
                output_wb.create_sheet("Sheet1")

            os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
            output_wb.save(output_path)
            logger.info(f"Formatted Excel file created: {output_path}")

            return output_path

        except Exception as e:
            logger.error(f"Error creating formatted Excel: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise

    def process_uploaded_file(self, uploaded_file_path):
        """
        Main method to process uploaded file

        Args:
            uploaded_file_path (str): Path to uploaded file

        Returns:
            str: Path to the generated output file
        """
        try:
            logger.info(f"Processing uploaded file: {uploaded_file_path}")

            file_type = self.detect_file_type(uploaded_file_path)
            logger.info(f"Detected file type: {file_type}")

            uploaded_data = self.read_file(uploaded_file_path)
            logger.info(f"Uploaded data sheets: {list(uploaded_data.keys())}")
            for sheet_name, data in uploaded_data.items():
                logger.info(f"Sheet '{sheet_name}' has {len(data)} rows")

            template_structure = self.analyze_template()

            logger.info("Starting OpenAI processing...")
            processed_data = self.process_with_openai(uploaded_data, template_structure, file_type)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"processed_{timestamp}.xlsx"
            output_path = os.path.join("document", output_filename)
            os.makedirs("document", exist_ok=True)

            logger.info("Creating formatted Excel file...")
            self.create_formatted_excel(processed_data, output_path)

            logger.info(f"File processing completed: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error processing uploaded file: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise

